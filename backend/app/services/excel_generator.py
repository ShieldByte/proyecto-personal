from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from io import BytesIO


# ── Colores Honda ────────────────────────────────────────────
ROJO    = "FFCC0000"
OSCURO  = "FF1a1a2e"
GRIS    = "FFF5F5F5"
VERDE   = "FF00a878"
NARANJA = "FFF4A261"
BLANCO  = "FFFFFFFF"
TEXTO   = "FF1a1a2e"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _borde():
    lado = Side(style="thin", color="FFE0E0E0")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _centrado(bold=False, size=10, color="FF000000"):
    return Font(bold=bold, size=size, color=color, name="Calibri")


def generar_excel(datos: dict) -> bytes:
    wb = Workbook()

    # ── HOJA 1: RESUMEN ──────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen del Turno"

    # Encabezado principal
    ws.merge_cells("A1:H1")
    ws["A1"] = "MES Honda Celaya — Reporte de Turno"
    ws["A1"].font = Font(bold=True, size=16, color=BLANCO, name="Calibri")
    ws["A1"].fill = _fill(OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Turno: {datos['turno']}   |   Fecha: {datos['fecha']}   |   Generado: {datos['fecha_generacion']}"
    ws["A2"].font = Font(size=10, color=BLANCO, name="Calibri")
    ws["A2"].fill = _fill(ROJO)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # KPIs — Producción
    ws["A4"] = "PRODUCCIÓN"
    ws["A4"].font = Font(bold=True, size=11, color=BLANCO, name="Calibri")
    ws["A4"].fill = _fill(OSCURO)
    ws.merge_cells("A4:D4")
    ws["A4"].alignment = Alignment(horizontal="center")

    kpis_prod = [
        ("Unidades Objetivo",    datos["totales"]["objetivo"]),
        ("Unidades Producidas",  datos["totales"]["producidas"]),
        ("Unidades Buenas",      datos["totales"]["buenas"]),
        ("Unidades Defectuosas", datos["totales"]["defectuosas"]),
        ("Cumplimiento",         f"{datos['totales']['cumplimiento']}%"),
    ]

    for i, (label, valor) in enumerate(kpis_prod, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = valor
        ws[f"A{i}"].font = Font(bold=True, size=10, name="Calibri", color=TEXTO)
        ws[f"B{i}"].font = Font(bold=True, size=10, name="Calibri", color=ROJO[2:])
        ws[f"A{i}"].fill = _fill(GRIS if i % 2 == 0 else BLANCO)
        ws[f"B{i}"].fill = _fill(GRIS if i % 2 == 0 else BLANCO)
        ws[f"A{i}"].border = _borde()
        ws[f"B{i}"].border = _borde()
        ws[f"B{i}"].alignment = Alignment(horizontal="center")

    # KPIs — OEE
    ws["E4"] = "OEE"
    ws["E4"].font = Font(bold=True, size=11, color=BLANCO, name="Calibri")
    ws["E4"].fill = _fill(OSCURO)
    ws.merge_cells("E4:H4")
    ws["E4"].alignment = Alignment(horizontal="center")

    oee = datos["oee"]
    kpis_oee = [
        ("OEE Global",      f"{oee['oee']}%"),
        ("Disponibilidad",  f"{oee['disponibilidad']}%"),
        ("Rendimiento",     f"{oee['rendimiento']}%"),
        ("Calidad",         f"{oee['calidad']}%"),
        ("Clasificación",   oee["clasificacion"]),
    ]

    for i, (label, valor) in enumerate(kpis_oee, start=5):
        ws[f"E{i}"] = label
        ws[f"F{i}"] = valor
        ws[f"E{i}"].font = Font(bold=True, size=10, name="Calibri", color=TEXTO)
        ws[f"F{i}"].font = Font(bold=True, size=10, name="Calibri", color=ROJO[2:])
        ws[f"E{i}"].fill = _fill(GRIS if i % 2 == 0 else BLANCO)
        ws[f"F{i}"].fill = _fill(GRIS if i % 2 == 0 else BLANCO)
        ws[f"E{i}"].border = _borde()
        ws[f"F{i}"].border = _borde()
        ws[f"F{i}"].alignment = Alignment(horizontal="center")

    # Ancho de columnas hoja resumen
    for col, ancho in zip("ABCDEFGH", [28, 16, 4, 4, 20, 16, 4, 4]):
        ws.column_dimensions[col].width = ancho

    # ── HOJA 2: ÓRDENES ─────────────────────────────────────
    ws2 = wb.create_sheet("Órdenes de Producción")

    encabezados = [
        "N° Orden", "Producto", "Estado", "Objetivo",
        "Producidas", "Buenas", "Defectuosas", "% Defectos", "Cumplimiento %"
    ]

    for col, enc in enumerate(encabezados, 1):
        cell = ws2.cell(row=1, column=col, value=enc)
        cell.font = Font(bold=True, size=10, color=BLANCO, name="Calibri")
        cell.fill = _fill(OSCURO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borde()

    ws2.row_dimensions[1].height = 24

    for fila, o in enumerate(datos["ordenes"], start=2):
        valores = [
            o["numero_orden"], o["producto"],
            o["estado"].replace("_", " ").title(),
            o["unidades_objetivo"], o["unidades_producidas"],
            o["unidades_buenas"], o["unidades_defectuosas"],
            o["pct_defectos"], o["cumplimiento"]
        ]
        bg = GRIS if fila % 2 == 0 else BLANCO

        for col, val in enumerate(valores, 1):
            cell = ws2.cell(row=fila, column=col, value=val)
            cell.font = Font(size=10, name="Calibri")
            cell.fill = _fill(bg)
            cell.border = _borde()
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left",
                                        vertical="center")

        # Colorear cumplimiento por semáforo
        c = o["cumplimiento"]
        color_c = VERDE if c >= 90 else (NARANJA if c >= 60 else ROJO)
        cell_c = ws2.cell(row=fila, column=9)
        cell_c.font = Font(bold=True, size=10, name="Calibri", color=color_c[2:])

    anchos2 = [18, 35, 16, 12, 14, 12, 14, 13, 15]
    for col, ancho in enumerate(anchos2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho

    # ── HOJA 3: ALERTAS ──────────────────────────────────────
    ws3 = wb.create_sheet("Alertas del Turno")

    enc_alertas = ["Hora", "Severidad", "Tipo", "Mensaje", "Estado"]
    for col, enc in enumerate(enc_alertas, 1):
        cell = ws3.cell(row=1, column=col, value=enc)
        cell.font = Font(bold=True, size=10, color=BLANCO, name="Calibri")
        cell.fill = _fill(OSCURO)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _borde()

    for fila, a in enumerate(datos["alertas"], start=2):
        colores_sev = {"critica": ROJO, "advertencia": NARANJA, "info": "FF4CC9F0"}
        bg = GRIS if fila % 2 == 0 else BLANCO
        valores_al = [
            a["hora"],
            a["severidad"].title(),
            a["tipo"].replace("_", " ").title(),
            a["mensaje"],
            "Resuelta" if a["resuelta"] else "Activa"
        ]
        for col, val in enumerate(valores_al, 1):
            cell = ws3.cell(row=fila, column=col, value=val)
            cell.font = Font(size=10, name="Calibri")
            cell.fill = _fill(bg)
            cell.border = _borde()

        # Color de severidad
        cell_sev = ws3.cell(row=fila, column=2)
        sev = a["severidad"]
        cell_sev.font = Font(
            bold=True, size=10, name="Calibri",
            color=colores_sev.get(sev, "FF000000")[2:]
        )

    anchos3 = [10, 16, 22, 55, 12]
    for col, ancho in enumerate(anchos3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = ancho

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()